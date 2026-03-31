import openpyxl
from openpyxl import Workbook
import os

source_file = "data/sample.xlsx"
output_folder = "data/"
os.makedirs(output_folder, exist_ok=True)

wb = openpyxl.load_workbook(source_file)

for sheet_name in wb.sheetnames:
    source_ws = wb[sheet_name]
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = sheet_name

    for row in source_ws.iter_rows():
        for cell in row:
            new_ws[cell.coordinate].value = cell.value

    new_wb.save(os.path.join(output_folder, f"{sheet_name}.xlsx"))